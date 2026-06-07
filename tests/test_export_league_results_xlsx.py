from pathlib import Path
import sys
from tempfile import TemporaryDirectory
import unittest

from openpyxl import Workbook, load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from exports.export_league_results_xlsx import (
    aggregate_league_results,
    build_workbook,
)


HEADERS = [
    "Data/Hora",
    "Jogo",
    "Competicao",
    "Tempo",
    "Placar",
    "Odds 1x2 Pre-live",
    "Link",
    "Saiu gol",
    "Unidade simples",
    "Saldo simples",
    "Unidade martingale",
    "Saldo martingale",
]


class ExportLeagueResultsXlsxTest(unittest.TestCase):
    def test_aggregates_leagues_across_all_sheets_sorted_by_accuracy(self):
        with TemporaryDirectory() as temp_dir:
            input_path = Path(temp_dir) / "input.xlsx"
            workbook = Workbook()
            first_sheet = workbook.active
            first_sheet.title = "Cenario 1"
            _write_sheet(first_sheet, [
                ("Brazil Serie A", "Sim"),
                ("Brazil Serie A", "Nao"),
                ("England Premier League", "Sim"),
            ])
            second_sheet = workbook.create_sheet("Cenario 2")
            _write_sheet(second_sheet, [
                ("Brazil Serie A", "Sim"),
                ("Spain La Liga", "Nao"),
            ])
            workbook.save(input_path)

            results = aggregate_league_results(str(input_path))

        self.assertEqual([item.league for item in results], [
            "England Premier League",
            "Brazil Serie A",
            "Spain La Liga",
        ])
        self.assertEqual((results[1].greens, results[1].reds), (2, 1))
        self.assertAlmostEqual(results[1].accuracy, 2 / 3)

    def test_builds_output_workbook_with_percentage_column(self):
        workbook = build_workbook(aggregate_like([
            ("League A", 3, 1),
            ("League B", 1, 1),
        ]))

        with TemporaryDirectory() as temp_dir:
            output_path = Path(temp_dir) / "output.xlsx"
            workbook.save(output_path)
            loaded = load_workbook(output_path, data_only=True)

        sheet = loaded["Resultados por Liga"]
        self.assertEqual(sheet["A1"].value, "Liga")
        self.assertEqual(sheet["E2"].value, 0.75)
        self.assertEqual(sheet["A4"].value, "TOTAL")
        self.assertEqual(sheet["E4"].value, 4 / 6)


def aggregate_like(items):
    from exports.export_league_results_xlsx import LeagueStats

    return [
        LeagueStats(league=league, greens=greens, reds=reds)
        for league, greens, reds in items
    ]


def _write_sheet(sheet, rows):
    sheet.append(["Cenario"])
    sheet.append(HEADERS)
    for league, outcome in rows:
        sheet.append([
            "2026-06-06 12:00:00",
            "Casa x Fora",
            league,
            "70",
            "1 x 0",
            "1.8 / 3.2 / 4",
            "https://cornerprobet.com/analysis/test",
            outcome,
            1,
            1,
            1,
            1,
        ])
    sheet.append(["TOTAL", "", "", "", "", "", "", "", 0, 0, 0, 0])


if __name__ == "__main__":
    unittest.main()
