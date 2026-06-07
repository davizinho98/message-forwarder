#!/usr/bin/env python3
"""
Aggregate league green/red results from mapa-de-calor scenario XLSX.
"""

from collections import defaultdict
from dataclasses import dataclass
import logging
from pathlib import Path
from typing import Dict, Iterable, List

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


INPUT_XLSX = "limite_76FT_por_cenario.xlsx"
OUTPUT_XLSX = "resultados_ligas_limite_76FT.xlsx"
HEADER_ROW = 2
DATA_START_ROW = 3
LEAGUE_HEADER = "Competicao"
OUTCOME_HEADER = "Saiu gol"


logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
)
logger = logging.getLogger(__name__)


@dataclass
class LeagueStats:
    league: str
    greens: int = 0
    reds: int = 0

    @property
    def total(self) -> int:
        return self.greens + self.reds

    @property
    def accuracy(self) -> float:
        if self.total == 0:
            return 0
        return self.greens / self.total


def aggregate_league_results(input_xlsx: str = INPUT_XLSX) -> List[LeagueStats]:
    workbook_path = Path(input_xlsx)
    if not workbook_path.exists():
        raise FileNotFoundError(f"Arquivo nao encontrado: {input_xlsx}")

    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    stats_by_league: Dict[str, LeagueStats] = defaultdict(lambda: LeagueStats(""))
    skipped_rows = 0

    for sheet in workbook.worksheets:
        header_indexes = _get_header_indexes(sheet)
        if LEAGUE_HEADER not in header_indexes or OUTCOME_HEADER not in header_indexes:
            logger.warning("Aba ignorada por cabecalho incompleto: %s", sheet.title)
            continue

        league_index = header_indexes[LEAGUE_HEADER] - 1
        outcome_index = header_indexes[OUTCOME_HEADER] - 1

        for row in sheet.iter_rows(min_row=DATA_START_ROW, values_only=True):
            first_cell = row[0]
            if _is_total_row(first_cell):
                break

            league = _normalize_text(row[league_index] if league_index < len(row) else None)
            outcome = _parse_outcome(row[outcome_index] if outcome_index < len(row) else None)

            if not league or outcome is None:
                skipped_rows += 1
                continue

            stats = stats_by_league[league]
            stats.league = league
            if outcome:
                stats.greens += 1
            else:
                stats.reds += 1

    results = sorted(
        stats_by_league.values(),
        key=lambda item: (-item.accuracy, -item.total, -item.greens, item.league.lower()),
    )

    logger.info("Ligas agregadas: %s", len(results))
    logger.info("Linhas ignoradas sem liga ou resultado valido: %s", skipped_rows)
    return results


def build_workbook(league_stats: Iterable[LeagueStats]) -> Workbook:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Resultados por Liga"

    headers = ["Liga", "Greens", "Reds", "Total", "Taxa de acerto"]
    sheet.append(headers)

    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
        cell.alignment = Alignment(horizontal="center")

    total_greens = 0
    total_reds = 0
    for stats in league_stats:
        total_greens += stats.greens
        total_reds += stats.reds
        sheet.append([
            stats.league,
            stats.greens,
            stats.reds,
            stats.total,
            stats.accuracy,
        ])
        sheet.cell(sheet.max_row, 5).number_format = "0.00%"

    total = total_greens + total_reds
    total_accuracy = total_greens / total if total else 0
    sheet.append(["TOTAL", total_greens, total_reds, total, total_accuracy])
    total_row = sheet.max_row
    for cell in sheet[total_row]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="FFF2CC")
    sheet.cell(total_row, 5).number_format = "0.00%"

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    _fit_columns(sheet)
    return workbook


def save_workbook(workbook: Workbook, output_xlsx: str = OUTPUT_XLSX) -> None:
    workbook.save(output_xlsx)


def export_league_results(
    input_xlsx: str = INPUT_XLSX,
    output_xlsx: str = OUTPUT_XLSX,
) -> None:
    league_stats = aggregate_league_results(input_xlsx)
    workbook = build_workbook(league_stats)
    save_workbook(workbook, output_xlsx)
    logger.info("XLSX gerado: %s", output_xlsx)


def _get_header_indexes(sheet) -> Dict[str, int]:
    return {
        str(cell.value).strip(): cell.column
        for cell in sheet[HEADER_ROW]
        if cell.value is not None
    }


def _parse_outcome(value) -> bool | None:
    normalized = _normalize_text(value).lower()
    if normalized in {"sim", "green", "true", "1"}:
        return True
    if normalized in {"nao", "no", "red", "false", "0"}:
        return False
    return None


def _normalize_text(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _is_total_row(value) -> bool:
    return _normalize_text(value).upper() == "TOTAL"


def _fit_columns(sheet) -> None:
    for column_cells in sheet.columns:
        column_letter = get_column_letter(column_cells[0].column)
        max_length = 0
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        sheet.column_dimensions[column_letter].width = min(max(max_length + 2, 10), 60)


def main() -> None:
    export_league_results()


if __name__ == "__main__":
    main()
