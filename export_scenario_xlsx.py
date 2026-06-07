#!/usr/bin/env python3
"""
Export mapa-de-calor Telegram alerts grouped by scenario to an XLSX file.
"""

import asyncio
from dataclasses import dataclass
from datetime import date, datetime, time, timezone
import json
import logging
import os
from pathlib import Path
import re
from typing import Dict, Iterable, List, Optional, Tuple, Union
from zoneinfo import ZoneInfo

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from scenario_classifier import SCENARIO_NAMES, AlertData, parse_and_classify


CONFIG_PATH = "client_config.json"
SOURCE_CHAT_ID = None
START_DATE = "2026-04-12"
END_DATE = "2026-06-06"
TIMEZONE = "America/Rio_Branco"
OUTPUT_XLSX = "mapa_de_calor_por_cenario2.xlsx"
SESSION_NAME = "scenario_xlsx_exporter"
STRATEGY_NAME = "mapa-de-calor"
CORNERPRO_BOT_ID = 779230055
CORNERPRO_BOT_USERNAME = "@cornerpro2_bot"
DIALOG_CACHE_LIMIT = 200


logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
)
logger = logging.getLogger(__name__)


HEADERS = [
    "Data/Hora",
    "Jogo",
    "Competicao",
    "Tempo",
    "Placar",
    "Odds 1x2 Pre-live",
    "Link",
    "Saiu gol",
    "Unidade simples",
    "Saldo simples",
    "Unidade martingale",
    "Saldo martingale",
]


@dataclass(frozen=True)
class ExportRow:
    message_datetime: datetime
    alert: AlertData
    scenario: str
    goal_scored: bool


@dataclass(frozen=True)
class UnitResult:
    simple_unit: int
    simple_balance: int
    martingale_unit: int
    martingale_balance: int


def load_config(config_path: str) -> dict:
    if os.getenv("API_ID"):
        config = {
            "api_id": int(os.getenv("API_ID")),
            "api_hash": os.getenv("API_HASH"),
        }
        if os.getenv("PHONE_NUMBER"):
            config["phone_number"] = os.getenv("PHONE_NUMBER")
        if os.getenv("BOT_TOKEN"):
            config["bot_token"] = os.getenv("BOT_TOKEN")
        if os.getenv("SCENARIO_SOURCE_CHAT_ID"):
            config["scenario_forwarders"] = [{
                "source_chat_id": int(os.getenv("SCENARIO_SOURCE_CHAT_ID")),
            }]
        return config

    with Path(config_path).open("r", encoding="utf-8") as file:
        return json.load(file)


def resolve_source_chat_id(config: dict) -> int:
    if SOURCE_CHAT_ID is not None:
        return int(SOURCE_CHAT_ID)

    forwarders = config.get("scenario_forwarders") or []
    if not forwarders:
        raise ValueError("Nenhum scenario_forwarders encontrado no config")

    source_chat_id = forwarders[0].get("source_chat_id")
    if source_chat_id is None:
        raise ValueError("scenario_forwarders[0].source_chat_id nao encontrado")
    return int(source_chat_id)


def validate_client_config(config: dict) -> None:
    for field in ["api_id", "api_hash"]:
        if not config.get(field):
            raise ValueError(f"Campo obrigatorio '{field}' nao encontrado na configuracao")

    if not config.get("phone_number") and not config.get("bot_token"):
        raise ValueError("Configure 'phone_number' ou 'bot_token' para conectar ao Telegram")


def parse_configured_date_range(
    start_date: str = START_DATE,
    end_date: str = END_DATE,
    timezone_name: str = TIMEZONE,
) -> Tuple[datetime, datetime]:
    if start_date == "YYYY-MM-DD" or end_date == "YYYY-MM-DD":
        raise ValueError("Edite START_DATE e END_DATE no arquivo antes de executar")

    tz = ZoneInfo(timezone_name)
    start = datetime.combine(date.fromisoformat(start_date), time.min, tzinfo=tz)
    end = datetime.combine(date.fromisoformat(end_date), time.max, tzinfo=tz)

    if end < start:
        raise ValueError("END_DATE precisa ser maior ou igual a START_DATE")

    return start, end


def message_datetime_to_local(message_datetime: datetime, timezone_name: str = TIMEZONE) -> datetime:
    if message_datetime.tzinfo is None:
        message_datetime = message_datetime.replace(tzinfo=timezone.utc)
    return message_datetime.astimezone(ZoneInfo(timezone_name))


def is_mapa_de_calor_strategy(strategy: str) -> bool:
    return _normalize_strategy(strategy) == _normalize_strategy(STRATEGY_NAME)


def extract_goal_outcome(message_text: str) -> Optional[bool]:
    goal_lines = []
    for raw_line in message_text.splitlines():
        line = raw_line.strip()
        if not line or "⚽" not in line:
            continue
        if re.search(r"\bresultado\s*:", line, flags=re.IGNORECASE):
            continue
        goal_lines.append(line)

    if not goal_lines:
        return None

    last_goal_line = goal_lines[-1]
    if "❌" in last_goal_line:
        return False
    return True


def calculate_unit_results(goal_results: Iterable[bool]) -> List[UnitResult]:
    results = []
    simple_balance = 0
    martingale_balance = 0
    consecutive_reds = 0
    pending_red_units = 0

    for goal_scored in goal_results:
        simple_unit = 1 if goal_scored else -1
        simple_balance += simple_unit

        stake = pending_red_units if consecutive_reds >= 2 else 1
        if goal_scored:
            martingale_unit = stake
            martingale_balance += stake
            consecutive_reds = 0
            pending_red_units = 0
        else:
            martingale_unit = -stake
            martingale_balance -= stake
            consecutive_reds += 1
            pending_red_units += stake

        results.append(UnitResult(
            simple_unit=simple_unit,
            simple_balance=simple_balance,
            martingale_unit=martingale_unit,
            martingale_balance=martingale_balance,
        ))

    return results


async def collect_rows(
    app,
    source_chat_id: Union[int, str],
    start_local: datetime,
    end_local: datetime,
) -> Dict[str, List[ExportRow]]:
    rows_by_scenario = {scenario: [] for scenario in SCENARIO_NAMES}
    skipped_without_result = 0
    skipped_unparseable = 0
    skipped_other_strategy = 0
    collected_descending = []

    async for message in app.get_chat_history(source_chat_id, limit=0):
        message_text = message.text or message.caption
        message_datetime = getattr(message, "date", None)
        if not message_datetime:
            continue

        local_datetime = message_datetime_to_local(message_datetime)
        if local_datetime > end_local:
            continue
        if local_datetime < start_local:
            break
        if not message_text:
            continue

        parsed = parse_and_classify(message_text)
        if not parsed:
            skipped_unparseable += 1
            continue

        alert, scenario_result = parsed
        if not is_mapa_de_calor_strategy(alert.strategy):
            skipped_other_strategy += 1
            continue

        goal_outcome = extract_goal_outcome(message_text)
        if goal_outcome is None:
            skipped_without_result += 1
            continue

        collected_descending.append(ExportRow(
            message_datetime=local_datetime,
            alert=alert,
            scenario=scenario_result.scenario,
            goal_scored=goal_outcome,
        ))

    for row in reversed(collected_descending):
        rows_by_scenario[row.scenario].append(row)

    logger.info("Mensagens exportaveis: %s", len(collected_descending))
    logger.info("Ignoradas sem resultado green/red claro: %s", skipped_without_result)
    logger.info("Ignoradas por formato nao parseavel: %s", skipped_unparseable)
    logger.info("Ignoradas por outra estrategia: %s", skipped_other_strategy)
    return rows_by_scenario


def build_workbook(rows_by_scenario: Dict[str, List[ExportRow]]) -> Workbook:
    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    sheet_titles = build_sheet_titles(SCENARIO_NAMES)
    for scenario in SCENARIO_NAMES:
        sheet = workbook.create_sheet(sheet_titles[scenario])
        rows = rows_by_scenario.get(scenario, [])
        _write_scenario_sheet(sheet, scenario, rows)

    return workbook


def build_sheet_titles(scenarios: Iterable[str]) -> Dict[str, str]:
    titles = {}
    used_titles = set()

    for index, scenario in enumerate(scenarios, 1):
        raw_title = f"{index:02d} {scenario}"
        title = re.sub(r"[\[\]\:\*\?/\\]", " ", raw_title).strip()[:31]
        title = title or f"Cenario {index:02d}"
        while title in used_titles:
            suffix = f" {index:02d}"
            title = f"{title[:31 - len(suffix)]}{suffix}"
        titles[scenario] = title
        used_titles.add(title)

    return titles


def save_workbook(workbook: Workbook, output_path: str) -> None:
    workbook.save(output_path)


async def export_scenario_xlsx() -> None:
    config = load_config(CONFIG_PATH)
    validate_client_config(config)
    source_chat_id = resolve_source_chat_id(config)
    start_local, end_local = parse_configured_date_range()

    logger.info("Exportando source_chat_id=%s", source_chat_id)
    logger.info("Range local: %s ate %s", start_local, end_local)

    client = _build_telegram_client(config)
    async with client as app:
        await warm_dialog_cache(app)
        source_peer = await resolve_source_peer(app, source_chat_id)
        rows_by_scenario = await collect_rows(app, source_peer, start_local, end_local)

    workbook = build_workbook(rows_by_scenario)
    save_workbook(workbook, OUTPUT_XLSX)
    logger.info("XLSX gerado: %s", OUTPUT_XLSX)


def _write_scenario_sheet(sheet, scenario: str, rows: List[ExportRow]) -> None:
    sheet.append([f"Cenario: {scenario}"])
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(HEADERS))
    sheet["A1"].font = Font(bold=True, size=12)
    sheet["A1"].fill = PatternFill("solid", fgColor="D9EAF7")

    sheet.append(HEADERS)
    for cell in sheet[2]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="F2F2F2")
        cell.alignment = Alignment(horizontal="center")

    unit_results = calculate_unit_results(row.goal_scored for row in rows)
    for row, unit_result in zip(rows, unit_results):
        alert = row.alert
        sheet.append([
            row.message_datetime.strftime("%Y-%m-%d %H:%M:%S"),
            f"{alert.home_team} x {alert.away_team}",
            alert.league,
            alert.game_time,
            f"{alert.home_goals} x {alert.away_goals}",
            f"{alert.home_odd:g} / {alert.draw_odd:g} / {alert.away_odd:g}",
            alert.match_url,
            "Sim" if row.goal_scored else "Nao",
            unit_result.simple_unit,
            unit_result.simple_balance,
            unit_result.martingale_unit,
            unit_result.martingale_balance,
        ])

    total_row_number = sheet.max_row + 1
    simple_total = unit_results[-1].simple_balance if unit_results else 0
    martingale_total = unit_results[-1].martingale_balance if unit_results else 0
    total_alerts = len(rows)
    total_hits = sum(1 for row in rows if row.goal_scored)
    hit_rate = total_hits / total_alerts if total_alerts else 0
    sheet.append([
        "TOTAL",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        simple_total,
        simple_total,
        martingale_total,
        martingale_total,
    ])
    for cell in sheet[total_row_number]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="FFF2CC")

    hit_rate_row_number = sheet.max_row + 1
    sheet.append([
        "Taxa de acerto",
        "",
        "",
        "",
        "",
        "",
        "",
        f"{total_hits}/{total_alerts}",
        hit_rate,
        "",
        "",
        "",
    ])
    sheet.cell(row=hit_rate_row_number, column=9).number_format = "0.00%"
    for cell in sheet[hit_rate_row_number]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="E2F0D9")

    sheet.freeze_panes = "A3"
    _fit_columns(sheet)


def _fit_columns(sheet) -> None:
    for column_cells in sheet.columns:
        column_letter = get_column_letter(column_cells[0].column)
        max_length = 0
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        sheet.column_dimensions[column_letter].width = min(max(max_length + 2, 10), 45)


def _build_telegram_client(config: dict):
    from pyrogram import Client

    client_kwargs = {
        "name": SESSION_NAME,
        "api_id": config["api_id"],
        "api_hash": config["api_hash"],
    }
    if config.get("phone_number"):
        client_kwargs["phone_number"] = config["phone_number"]
    else:
        logger.warning("Usando bot_token. Para historico completo, prefira phone_number.")
        client_kwargs["bot_token"] = config["bot_token"]

    return Client(**client_kwargs)


async def warm_dialog_cache(app, limit: int = DIALOG_CACHE_LIMIT) -> None:
    dialog_count = 0
    try:
        async for _dialog in app.get_dialogs(limit=limit):
            dialog_count += 1
        logger.info("Cache de dialogos carregado com %s dialogos", dialog_count)
    except Exception as error:
        logger.warning("Nao foi possivel aquecer o cache de dialogos: %s", error)


async def resolve_source_peer(app, source_chat_id: int) -> Union[int, str]:
    if source_chat_id == CORNERPRO_BOT_ID:
        try:
            user = await app.get_users(CORNERPRO_BOT_USERNAME)
            logger.info(
                "Fonte resolvida por username: %s (@%s, ID %s)",
                user.first_name,
                user.username or "sem_username",
                user.id,
            )
            return CORNERPRO_BOT_USERNAME
        except Exception as error:
            logger.warning(
                "Nao foi possivel resolver %s por username: %s",
                CORNERPRO_BOT_USERNAME,
                error,
            )

    try:
        chat = await app.get_chat(source_chat_id)
        source_name = getattr(chat, "title", None) or getattr(chat, "first_name", source_chat_id)
        logger.info("Fonte resolvida por ID: %s (%s)", source_name, source_chat_id)
        return chat.id
    except Exception as first_error:
        logger.warning("Falha ao resolver fonte %s por ID: %s", source_chat_id, first_error)

    await warm_dialog_cache(app)

    try:
        chat = await app.get_chat(source_chat_id)
        source_name = getattr(chat, "title", None) or getattr(chat, "first_name", source_chat_id)
        logger.info("Fonte resolvida apos cache: %s (%s)", source_name, source_chat_id)
        return chat.id
    except Exception as second_error:
        raise RuntimeError(
            f"Nao foi possivel resolver a fonte {source_chat_id}. "
            "Se for o CornerProBot2, envie /start para @cornerpro2_bot e rode novamente. "
            "Se for grupo/canal, confirme que o usuario da sessao participa dele e que o ID esta correto."
        ) from second_error


def _normalize_strategy(strategy: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", strategy.lower())


def main() -> None:
    asyncio.run(export_scenario_xlsx())


if __name__ == "__main__":
    main()
